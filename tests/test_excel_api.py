import io
import json
import unittest

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Border, Font, PatternFill, Side

from app import app


def make_styled_workbook_bytes():
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "2023"

    target = worksheet["F5"]
    target.value = 100
    target.font = Font(name="Calibri", size=14, bold=True, color="00FF0000")
    target.fill = PatternFill(fill_type="solid", fgColor="00FFF2CC")
    thin = Side(style="thin", color="00000000")
    target.border = Border(left=thin, right=thin, top=thin, bottom=thin)

    buffer = io.BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer


class ExcelApiTestCase(unittest.TestCase):
    def setUp(self):
        app.config["TESTING"] = True
        self.client = app.test_client()

    def test_apply_updates_preserves_cell_style(self):
        response = self.client.post(
            "/api/excel/apply-updates",
            data={
                "file": (make_styled_workbook_bytes(), "dien_so_dien.xlsx"),
                "updates": json.dumps([
                    {"sheet": "2023", "address": "F5", "value": 12345},
                ]),
                "filename": "KetQua_Excel.xlsx",
            },
            content_type="multipart/form-data",
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response.headers["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        self.assertIn("KetQua_Excel.xlsx", response.headers["Content-Disposition"])

        output_wb = load_workbook(io.BytesIO(response.data))
        output_cell = output_wb["2023"]["F5"]

        self.assertEqual(output_cell.value, 12345)
        self.assertTrue(output_cell.font.bold)
        self.assertEqual(output_cell.font.size, 14)
        self.assertEqual((output_cell.fill.fgColor.rgb or "").upper(), "00FFF2CC")
        self.assertEqual(output_cell.border.left.style, "thin")
        self.assertEqual(output_cell.border.right.style, "thin")
        self.assertEqual(output_cell.border.top.style, "thin")
        self.assertEqual(output_cell.border.bottom.style, "thin")

    def test_apply_updates_rejects_unknown_sheet(self):
        response = self.client.post(
            "/api/excel/apply-updates",
            data={
                "file": (make_styled_workbook_bytes(), "dien_so_dien.xlsx"),
                "updates": json.dumps([
                    {"sheet": "2024", "address": "F5", "value": 12345},
                ]),
            },
            content_type="multipart/form-data",
        )

        self.assertEqual(response.status_code, 400)
        self.assertEqual(
            response.get_json()["error"],
            "Không tìm thấy sheet '2024' trong file Excel.",
        )


if __name__ == "__main__":
    unittest.main()
