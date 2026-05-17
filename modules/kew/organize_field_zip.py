"""Mô-đun hỗ trợ tổ chức hồ sơ đo KEW6315 từ file ZIP.

Mô-đun này thực hiện các công việc:
1. Đọc file Excel hiện trường (.xlsx) để lấy kế hoạch.
2. Đổi tên các thư mục Sxxxx thành tên thiết bị tương ứng.
3. Di chuyển các file ảnh PS-SDxxx.BMP vào đúng thư mục thiết bị.
4. Chạy OCR tự động đọc các thông số đo từ ảnh và ghi vào Excel.
5. Nén lại thư mục kết quả thành Project_Output.zip.
"""
from __future__ import annotations

import io
import os
import re
import shutil
import unicodedata
import zipfile
from dataclasses import dataclass
from typing import Any, Optional

import pandas as pd

_SKIP_DIR_NAMES = {"__MACOSX"}

# ── Cột bắt buộc người dùng phải điền vào Excel hiện trường ──────────────────
FIELD_XLSX_REQUIRED: tuple[str, ...] = (
    "stt",           # Số thứ tự (thứ tự trong báo cáo Word)
    "name",          # Tên thiết bị (hiển thị trong báo cáo)
    "file",          # Mã thư mục KEW (Sxxxx)
    "img",           # Chỉ số ảnh đầu dải (PS-SDxxx)
    "imgend",        # Chỉ số ảnh cuối dải
    "type",          # Loại section: MBA / device (thiết bị)
    "pdm",           # Công suất định mức (kVA) — dùng tính % tải MBA
    "current_char",  # Đặc tính dòng điện: Ổn định / Dao động nhẹ / ...
)

# ── Cột tự chọn (không bắt buộc có trong Excel) ──────────────────────────────
FIELD_XLSX_OPTIONAL: tuple[str, ...] = (
    "imgomit",       # Chỉ số ảnh bỏ qua trong dải (tuỳ chọn)
    "imglu",         # Chỉ số ảnh load/unload → đổi tên load-unload-xxx.BMP
)

# ── Cột do OCR tự động điền sau khi nhận dạng ảnh BMP ────────────────────────
FIELD_XLSX_OCR: tuple[str, ...] = (
    "p",             # Công suất tác dụng trung bình (kW)
    "cos_phi",       # Hệ số công suất trung bình (cosφ)
    "i_max",         # Dòng điện lớn nhất đo được (A)
    "u_min",         # Điện áp đo thấp nhất (V)
    "u_max",         # Điện áp đo cao nhất (V)
    "delta_u",       # Mất cân bằng điện áp lớn nhất (%)
    "delta_i",       # Mất cân bằng dòng điện lớn nhất (%)
    "thd",           # THD điện áp lớn nhất (%)
    "tdd",           # TDD dòng điện lớn nhất (%)
)

# ── Toàn bộ schema đầy đủ — dùng cho Word report ─────────────────────────────
FIELD_XLSX_HEADERS: tuple[str, ...] = FIELD_XLSX_REQUIRED + FIELD_XLSX_OPTIONAL + FIELD_XLSX_OCR

_BMP_RE = re.compile(r"^PS-SD(\d{1,4})\.BMP$", re.IGNORECASE)
_S_DIR_RE = re.compile(r"^S(\d{4})$", re.IGNORECASE)

_WIN_RESERVED = {
    "CON", "PRN", "AUX", "NUL",
    *(f"COM{i}" for i in range(1, 10)),
    *(f"LPT{i}" for i in range(1, 10)),
}


def _norm_key(s: str) -> str:
    """Chuẩn hóa chuỗi để so khớp tên cột.

    Chuyển đổi chuỗi sang dạng NFKC, viết thường và loại bỏ khoảng trắng thừa.

    Args:
        s (str): Chuỗi cần chuẩn hóa.

    Returns:
        str: Chuỗi đã được chuẩn hóa.
    """
    if s is None or (isinstance(s, float) and pd.isna(s)):
        return ""
    t = unicodedata.normalize("NFKC", str(s)).strip().lower()
    t = re.sub(r"\s+", " ", t)
    return t


def _is_skipped_path(path: str) -> bool:
    """Kiểm tra xem đường dẫn có nên bị bỏ qua hay không.

    Bỏ qua các thư mục hệ thống như __MACOSX và các file ẩn bắt đầu bằng ._.

    Args:
        path (str): Đường dẫn cần kiểm tra.

    Returns:
        bool: True nếu nên bỏ qua, ngược lại False.
    """
    parts = path.split(os.sep)
    return any(p in _SKIP_DIR_NAMES or p.startswith("._") for p in parts)


def find_first_excel(root: str) -> Optional[str]:
    """Tìm file Excel đầu tiên trong thư mục gốc.

    Quét đệ quy thư mục gốc để tìm file có đuôi .xlsx hoặc .xlsm.

    Args:
        root (str): Thư mục để tìm kiếm.

    Returns:
        Optional[str]: Đường dẫn đến file Excel tìm được hoặc None nếu không thấy.
    """
    candidates: list[str] = []
    for dirpath, _, filenames in os.walk(root):
        if _is_skipped_path(dirpath):
            continue
        for fn in filenames:
            low = fn.lower()
            if low.endswith((".xlsx", ".xlsm")):
                candidates.append(os.path.join(dirpath, fn))
    if not candidates:
        return None
    candidates.sort()
    return candidates[0]


def scan_s_folders(root: str) -> tuple[dict[str, str], list[str]]:
    """Quét tất cả các thư mục có định dạng Sxxxx.

    Tìm kiếm các thư mục có tên khớp với định dạng Sxxxx (ví dụ: S0001) trong thư mục gốc.

    Args:
        root (str): Thư mục gốc cần quét.

    Returns:
        tuple[dict[str, str], list[str]]: Một tuple gồm:
            - Dict mapping mã thư mục (Sxxxx) sang đường dẫn tuyệt đối.
            - Danh sách các thông báo lỗi nếu phát hiện trùng lặp.
    """
    mapping: dict[str, str] = {}
    errors: list[str] = []
    for dirpath, dirnames, _ in os.walk(root):
        if _is_skipped_path(dirpath):
            continue
        for d in dirnames:
            m = _S_DIR_RE.match(d)
            if not m:
                continue
            key = f"S{m.group(1)}"
            full = os.path.join(dirpath, d)
            if key in mapping and os.path.normpath(mapping[key]) != os.path.normpath(full):
                errors.append(f"Trùng thư mục {key}: {mapping[key]} và {full}")
            else:
                mapping[key] = full
    return mapping, errors


def scan_bmp_files(root: str) -> tuple[dict[int, str], list[str]]:
    """Quét tất cả các file ảnh BMP có định dạng PS-SDxxxx.BMP.

    Trích xuất số thứ tự ảnh từ tên file và lưu vào map.

    Args:
        root (str): Thư mục gốc cần quét.

    Returns:
        tuple[dict[int, str], list[str]]: Một tuple gồm:
            - Dict mapping số thứ tự ảnh (int) sang đường dẫn tuyệt đối.
            - Danh sách các thông báo lỗi nếu phát hiện trùng lặp số ảnh.
    """
    by_num: dict[int, str] = {}
    dup: list[str] = []
    for dirpath, _, filenames in os.walk(root):
        if _is_skipped_path(dirpath):
            continue
        for fn in filenames:
            m = _BMP_RE.match(fn)
            if not m:
                continue
            n = int(m.group(1))
            full = os.path.join(dirpath, fn)
            if n in by_num and os.path.normpath(by_num[n]) != os.path.normpath(full):
                dup.append(f"Trùng PS-SD{n:03d}: {by_num[n]} và {full}")
            else:
                by_num[n] = full
    return by_num, dup


def file_code_to_s_name(raw: Any) -> Optional[str]:
    """Chuẩn hóa mã file từ Excel thành chuỗi chuẩn 'Sxxxx'.

    Ví dụ: '447' hoặc 'S0447' hoặc 447 đều chuyển thành 'S0447'.

    Args:
        raw (Any): Giá trị thô từ ô Excel.

    Returns:
        Optional[str]: Chuỗi 'Sxxxx' hoặc None nếu không hợp lệ.
    """
    if raw is None or (isinstance(raw, float) and pd.isna(raw)):
        return None
    s = str(raw).strip()
    if not s:
        return None
    m = re.match(r"^S\s*(\d{1,4})$", s, re.I)
    if m:
        return f"S{int(m.group(1)):04d}"
    digits = re.sub(r"\D", "", s)
    if not digits:
        return None
    return f"S{int(digits):04d}"


def _to_int_img(v: Any) -> Optional[int]:
    """Chuyển đổi chỉ số ảnh sang kiểu số nguyên.

    Hỗ trợ chuyển đổi từ chuỗi, số thực có giá trị nguyên.

    Args:
        v (Any): Giá trị cần chuyển đổi.

    Returns:
        Optional[int]: Giá trị số nguyên hoặc None nếu không hợp lệ.
    """
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return None
    if isinstance(v, (int, float)) and not isinstance(v, bool):
        if isinstance(v, float) and v.is_integer():
            return int(v)
        if isinstance(v, int):
            return int(v)
    s = str(v).strip()
    if not s:
        return None
    digits = re.sub(r"\D", "", s)
    if not digits:
        return None
    return int(digits)


def _parse_img_omit(raw: Any) -> tuple[frozenset[int], list[str]]:
    """Phân tích danh sách chỉ số ảnh cần bỏ qua từ cột imgomit.

    Hỗ trợ các định dạng: "944", "944,945", "944+945", "PS-SD944", "PS-SD944.BMP".

    Args:
        raw (Any): Giá trị thô từ ô Excel.

    Returns:
        tuple[frozenset[int], list[str]]: Một tuple gồm:
            - Set chứa các chỉ số ảnh cần bỏ qua (frozenset).
            - Danh sách các cảnh báo nếu có phần tử không hợp lệ.
    """
    warnings: list[str] = []
    if raw is None or (isinstance(raw, float) and pd.isna(raw)):
        return frozenset(), warnings
    if isinstance(raw, (int, float)) and not isinstance(raw, bool):
        if isinstance(raw, float):
            if not raw.is_integer():
                warnings.append("imgomit: ô chứa số thập phân — chỉ dùng chỉ số ảnh nguyên, bỏ qua giá trị này.")
                return frozenset(), warnings
            return frozenset({int(raw)}), warnings
        return frozenset({raw}), warnings
    s = str(raw).strip()
    if not s:
        return frozenset(), warnings
    out: set[int] = set()
    for tok in re.split(r"[,;+/\s|]+", s):
        tok = tok.strip()
        if not tok:
            continue
        tm = re.search(r"PS-SD(\d{1,4})(?:\.BMP)?", tok, re.IGNORECASE)
        if tm:
            out.add(int(tm.group(1)))
            continue
        dm = re.fullmatch(r"(\d{1,4})", re.sub(r"\s+", "", tok))
        if dm:
            out.add(int(dm.group(1)))
            continue
        warnings.append(f"imgomit: không hiểu mục «{tok}» — bỏ qua token.")
    return frozenset(out), warnings


def bmp_basename_for_index(n: int) -> str:
    """Tạo tên file BMP từ chỉ số ảnh.

    Ví dụ: 1 -> "PS-SD001.BMP".

    Args:
        n (int): Chỉ số ảnh.

    Returns:
        str: Tên file BMP tương ứng.
    """
    return f"PS-SD{n:03d}.BMP"


def sanitize_device_folder(name: Any) -> str:
    """Làm sạch tên thiết bị để dùng làm tên thư mục.

    Xóa các ký tự cấm trong tên file của Windows/Linux, chuẩn hóa về dạng NFC.

    Args:
        name (Any): Tên thiết bị thô.

    Returns:
        str: Tên thư mục đã được làm sạch.

    Raises:
        ValueError: Nếu tên thiết bị trống hoặc không hợp lệ sau chuẩn hóa.
    """
    if name is None or (isinstance(name, float) and pd.isna(name)):
        raise ValueError("Tên thiết bị trống.")
    s = unicodedata.normalize("NFKC", str(name)).strip()
    s = re.sub(r'[<>:"/\\|?*\x00-\x1f]', "_", s)
    s = re.sub(r"\s+", " ", s).strip()
    s = s.rstrip(" .")
    if not s or s in {".", ".."}:
        raise ValueError("Tên thiết bị không hợp lệ sau khi chuẩn hóa.")
    base = s.split(".")[0].upper()
    if base in _WIN_RESERVED:
        s = f"_{s}"
    if len(s) > 180:
        s = s[:180].rstrip(" .")
    # Dùng NFC để tên thư mục trùng khóa với Excel / macOS (HFS+ hay lưu NFD).
    return unicodedata.normalize("NFC", s)


def _unique_name(base: str, used: set[str]) -> str:
    """Đảm bảo tên thư mục là duy nhất trong tập hợp đã dùng.

    Nếu tên đã tồn tại, tự động thêm hậu tố _2, _3...

    Args:
        base (str): Tên thư mục cơ sở.
        used (set[str]): Tập hợp các tên đã sử dụng.

    Returns:
        str: Tên thư mục duy nhất.
    """
    if base not in used:
        used.add(base)
        return base
    i = 2
    while True:
        cand = f"{base}_{i}"
        if cand not in used:
            used.add(cand)
            return cand
        i += 1


def resolve_field_excel_column_map(df: pd.DataFrame) -> dict[str, Any]:
    """Map tên cột logic sang tên cột gốc trong DataFrame.

    Kiểm tra các cột bắt buộc. Các cột tùy chọn hoặc OCR nếu thiếu sẽ được map với None.

    Args:
        df (pd.DataFrame): DataFrame đọc từ file Excel.

    Returns:
        dict[str, Any]: Dict mapping tên cột logic sang tên cột gốc trong file Excel.

    Raises:
        ValueError: Nếu thiếu cột bắt buộc hoặc có 2 cột trùng tên sau chuẩn hóa.
    """
    seen: dict[str, Any] = {}
    for c in df.columns:
        k = _norm_key(str(c))
        if not k:
            continue
        if k in seen and seen[k] != c:
            raise ValueError(
                f"Hai cột trùng tên sau khi chuẩn hóa ({k!r}): {seen[k]!r} và {c!r}."
            )
        if k not in seen:
            seen[k] = c

    # Cột không nhận biết: chỉ cảnh báo, không lỗi (ví dụ cột ghi chú riêng của user)
    all_known = set(FIELD_XLSX_HEADERS)
    # (không raise — bỏ qua cột không biết)

    # Kiểm tra các cột BẮT BUỘC
    missing_required = [h for h in FIELD_XLSX_REQUIRED if h not in seen]
    if missing_required:
        raise ValueError(
            "File Excel hiện trường thiếu các cột bắt buộc: "
            + ", ".join(missing_required)
            + ". Cột bắt buộc: "
            + ", ".join(FIELD_XLSX_REQUIRED)
            + f". Các cột hiện có: {list(df.columns)}"
        )

    # Trả về map đầy đủ: cột vắng mặt (OCR) → None
    result: dict[str, Any] = {}
    for h in FIELD_XLSX_HEADERS:
        result[h] = seen.get(h)  # None nếu không có
    return result


@dataclass
class RowPlan:
    device_raw: str
    folder_name: str
    s_key: str
    img_start: int
    img_end: int
    img_omit: frozenset[int]
    img_lu: Optional[int]  # Chỉ số ảnh load/unload (None nếu không có)
    excel_row: int         # Chỉ số hàng 1-based trong Excel (bao gồm header)


def read_plans_from_excel(excel_path: str) -> tuple[list[RowPlan], list[str]]:
    """Đọc kế hoạch tổ chức hồ sơ từ file Excel hiện trường.

    Phân tích các dòng trong file Excel để tạo danh sách kế hoạch xử lý.

    Args:
        excel_path (str): Đường dẫn đến file Excel.

    Returns:
        tuple[list[RowPlan], list[str]]: Một tuple gồm:
            - Danh sách các đối tượng RowPlan chứa thông tin kế hoạch.
            - Danh sách các cảnh báo phát sinh trong quá trình đọc.

    Raises:
        ValueError: Nếu file Excel rỗng hoặc không có dòng hợp lệ nào.
    """
    warnings: list[str] = []
    try:
        df = pd.read_excel(excel_path, header=0, engine="openpyxl")
    except Exception as e:
        raise ValueError(f"Không đọc được Excel: {e}") from e
    if df.empty:
        raise ValueError("File Excel không có dữ liệu.")
    colmap = resolve_field_excel_column_map(df)
    used_names: set[str] = set()
    plans: list[RowPlan] = []
    used_s: set[str] = set()

    for idx, row in df.iterrows():
        try:
            dev_raw = row[colmap["name"]]
            s_name = file_code_to_s_name(row[colmap["file"]])
            i0 = _to_int_img(row[colmap["img"]])
            i1 = _to_int_img(row[colmap["imgend"]])
        except Exception as e:
            warnings.append(f"Dòng {int(idx) + 2}: bỏ qua ({e})")
            continue
        if s_name is None or i0 is None or i1 is None:
            warnings.append(f"Dòng {int(idx) + 2}: thiếu file/img/imgend — bỏ qua.")
            continue
        if i1 < i0:
            raise ValueError(f"Dòng {int(idx) + 2}: IMG end ({i1}) nhỏ hơn IMG ({i0}).")
        if s_name in used_s:
            raise ValueError(f"Dòng {int(idx) + 2}: mã thư mục {s_name} bị lặp trong Excel.")
        used_s.add(s_name)
        omit_col = colmap.get("imgomit")
        omit_all, w_omit = _parse_img_omit(row[omit_col]) if omit_col else (frozenset(), [])
        warnings.extend(w_omit)
        omit_eff = frozenset(n for n in omit_all if i0 <= n <= i1)
        outs = sorted(n for n in omit_all if n < i0 or n > i1)
        if outs:
            warnings.append(
                f"Dòng {int(idx) + 2}: imgomit có chỉ số ngoài dải {i0}–{i1}: "
                f"{', '.join(str(x) for x in outs)} (bỏ qua các mục ngoài dải)."
            )
        # Cột imglu: tuỳ chọn — ảnh load/unload dạng sóng (ảnh thứ 7)
        lu_col = colmap.get("imglu")
        img_lu = _to_int_img(row[lu_col]) if lu_col else None
        try:
            folder = sanitize_device_folder(dev_raw)
        except ValueError as e:
            raise ValueError(f"Dòng {int(idx) + 2}: {e}") from e
        folder = _unique_name(folder, used_names)
        plans.append(
            RowPlan(
                device_raw=str(dev_raw).strip(),
                folder_name=folder,
                s_key=s_name,
                img_start=i0,
                img_end=i1,
                img_omit=omit_eff,
                img_lu=img_lu,
                excel_row=int(idx) + 2,  # idx 0-based của data, row 1-indexed (header=row 1)
            )
        )
    if not plans:
        raise ValueError("Không có dòng hợp lệ nào trong Excel (sau khi lọc).")
    return plans, warnings

def _ranges_overlap(a0: int, a1: int, b0: int, b1: int) -> bool:
    """Kiểm tra xem hai dải ảnh có bị chồng lấn hay không.

    Args:
        a0 (int): Điểm bắt đầu dải A.
        a1 (int): Điểm kết thúc dải A.
        b0 (int): Điểm bắt đầu dải B.
        b1 (int): Điểm kết thúc dải B.

    Returns:
        bool: True nếu có chồng lấn, ngược lại False.
    """
    return not (a1 < b0 or b1 < a0)


def validate_plans_against_fs(
    plans: list[RowPlan],
    s_map: dict[str, str],
    bmp_map: dict[int, str],
) -> list[str]:
    """Kiểm tra tính hợp lệ của kế hoạch so với file hệ thống.

    Kiểm tra xem các thư mục Sxxxx và file ảnh BMP có tồn tại hay không,
    và kiểm tra trùng lặp dải ảnh giữa các thiết bị.

    Args:
        plans (list[RowPlan]): Danh sách kế hoạch.
        s_map (dict[str, str]): Map thư mục Sxxxx hiện có.
        bmp_map (dict[int, str]): Map file ảnh BMP hiện có.

    Returns:
        list[str]: Danh sách các thông báo lỗi (nếu có).
    """
    errors: list[str] = []
    for p in plans:
        if p.s_key not in s_map:
            errors.append(f"Thiết bị «{p.device_raw}»: không có thư mục {p.s_key} trong ZIP.")
        kept = [n for n in range(p.img_start, p.img_end + 1) if n not in p.img_omit]
        if not kept:
            errors.append(
                f"Thiết bị «{p.device_raw}»: imgomit loại hết dải ảnh {p.img_start}–{p.img_end}."
            )
        for n in kept:
            if n not in bmp_map:
                errors.append(
                    f"Thiết bị «{p.device_raw}»: thiếu ảnh {bmp_basename_for_index(n)}."
                )
    for i, pa in enumerate(plans):
        for pb in plans[i + 1 :]:
            if _ranges_overlap(pa.img_start, pa.img_end, pb.img_start, pb.img_end):
                errors.append(
                    f"Trùng dải ảnh: «{pa.device_raw}» ({pa.img_start}–{pa.img_end}) "
                    f"và «{pb.device_raw}» ({pb.img_start}–{pb.img_end})."
                )
    return errors


def build_project_output(
    extract_root: str,
    output_parent: str,
    plans: list[RowPlan],
    s_map: dict[str, str],
    bmp_map: dict[int, str],
    warnings: Optional[list[str]] = None,
) -> str:
    """Tạo thư mục Project_Output và copy file theo kế hoạch.

    Args:
        extract_root (str): Thư mục giải nén ZIP gốc.
        output_parent (str): Thư mục chứa thư mục Project_Output kết quả.
        plans (list[RowPlan]): Danh sách kế hoạch.
        s_map (dict[str, str]): Map thư mục Sxxxx.
        bmp_map (dict[int, str]): Map file ảnh BMP.
        warnings (Optional[list[str]]): Danh sách để lưu cảnh báo.

    Returns:
        str: Đường dẫn đến thư mục Project_Output đã tạo.
    """
    if warnings is None:
        warnings = []
    out_root = os.path.join(output_parent, "Project_Output")
    os.makedirs(out_root, exist_ok=True)
    for p in plans:
        src_dir = s_map[p.s_key]
        dest_dir = os.path.join(out_root, p.folder_name)
        if os.path.exists(dest_dir):
            shutil.rmtree(dest_dir)
        shutil.copytree(src_dir, dest_dir)

        for n in range(p.img_start, p.img_end + 1):
            if n in p.img_omit:
                continue
            src_bmp = bmp_map[n]
            norm_src = os.path.normpath(src_bmp)
            base = bmp_basename_for_index(n)
            dest_bmp = os.path.join(dest_dir, base)

            if os.path.normpath(dest_bmp) == norm_src:
                continue

            if os.path.dirname(norm_src) == os.path.normpath(dest_dir):
                continue

            shutil.copy2(src_bmp, dest_bmp)
            try:
                os.remove(src_bmp)
            except OSError:
                pass

        # ── Ảnh load/unload (imglu) — tuỳ chọn, ảnh thứ 7 ──────────────
        if p.img_lu is not None:
            if p.img_lu not in bmp_map:
                warnings.append(
                    f"Thiết bị «{p.device_raw}»: không tìm thấy ảnh "
                    f"{bmp_basename_for_index(p.img_lu)} cho imglu — bỏ qua."
                )
            else:
                lu_dest_name = f"load-unload-{p.img_lu:03d}.BMP"
                lu_dest = os.path.join(dest_dir, lu_dest_name)
                # Nếu img_lu nằm trong dải thường, file đã được copy vào dest_dir
                # với tên PS-SDxxx.BMP (và nguồn gốc có thể đã bị xoá).
                # Ưu tiên dùng bản đã copy; chỉ fallback sang bmp_map nếu chưa có.
                already_copied = os.path.join(dest_dir, bmp_basename_for_index(p.img_lu))
                if os.path.exists(already_copied):
                    lu_src = already_copied
                    shutil.copy2(lu_src, lu_dest)
                    # Không xoá bản gốc vì nó đã nằm trong dest_dir
                else:
                    lu_src = bmp_map[p.img_lu]
                    shutil.copy2(lu_src, lu_dest)
                    # Xoá nguồn nếu nằm ngoài thư mục đích
                    if os.path.dirname(os.path.normpath(lu_src)) != os.path.normpath(dest_dir):
                        try:
                            os.remove(lu_src)
                        except OSError:
                            pass

    return out_root


def zip_directory(folder: str, zip_path: str) -> None:
    """Nén một thư mục thành file ZIP.

    Args:
        folder (str): Thư mục cần nén.
        zip_path (str): Đường dẫn file ZIP đầu ra.
    """
    parent = os.path.dirname(folder)
    with zipfile.ZipFile(zip_path, "w", compression=zipfile.ZIP_DEFLATED) as zf:
        for root, _, files in os.walk(folder):
            for fn in files:
                fp = os.path.join(root, fn)
                arc = os.path.relpath(fp, parent)
                zf.write(fp, arcname=arc)


def run_ocr_and_update_excel(
    excel_path: str,
    plans: list[RowPlan],
    bmp_map: dict[int, str],
    overwrite_existing: bool = False,
) -> list[str]:
    """Chạy OCR nhận dạng thông số từ ảnh và cập nhật vào file Excel.

    Args:
        excel_path (str): Đường dẫn đến file Excel cần cập nhật.
        plans (list[RowPlan]): Danh sách kế hoạch.
        bmp_map (dict[int, str]): Map file ảnh BMP.
        overwrite_existing (bool): Có ghi đè các ô đã có dữ liệu hay không.

    Returns:
        list[str]: Danh sách các cảnh báo nhận dạng được.
    """
    try:
        from openpyxl import load_workbook
    except ImportError:
        return ["Thiếu thư viện openpyxl — bỏ qua bước OCR điền Excel."]

    try:
        from modules.image.ocr_kew import read_device_ocr
    except ImportError as e:
        return [f"Không tải được module OCR ({e}) — bỏ qua bước OCR."]

    warnings_out: list[str] = []

    # Mở workbook để ghi trực tiếp
    try:
        wb = load_workbook(excel_path)
        ws = wb.active
    except Exception as e:
        return [f"OCR: không mở được workbook để ghi ({e})."]

    header_row = 1

    # ── Quét header hiện có trong worksheet ──────────────────────────────────
    existing_col_map: dict[str, int] = {}  # cột chuẩn hóa → col_idx (1-indexed)
    max_col = 0
    for col_idx, cell in enumerate(ws[header_row], start=1):
        max_col = max(max_col, col_idx)
        if cell.value is not None:
            k = _norm_key(str(cell.value))
            if k:
                existing_col_map[k] = col_idx

    # ── Đảm bảo mỗi cột OCR tồn tại trong header; tạo mới nếu chưa có ───────
    col_indices: dict[str, int] = {}
    for field in FIELD_XLSX_OCR:
        if field in existing_col_map:
            col_indices[field] = existing_col_map[field]
        else:
            max_col += 1
            ws.cell(row=header_row, column=max_col, value=field)
            col_indices[field] = max_col
            warnings_out.append(f"OCR: tạo mới cột «{field}» tại cột số {max_col}.")

    # Tìm cột 'type' để lọc MBA (dựa trên header thực tế)
    type_col_idx = existing_col_map.get("type")

    for plan in plans:
        # Bỏ qua MBA nếu không có thông tin ảnh (trước đây bỏ qua cứng, nay cho phép nếu có img_start)
        if type_col_idx:
            row_type_val = ws.cell(row=plan.excel_row, column=type_col_idx).value
            row_type = str(row_type_val).strip().lower() if row_type_val else ""
            # Vẫn cho phép OCR MBA nếu user có điền chỉ số ảnh
            if not plan.img_start:
                continue

        # Lấy danh sách ảnh thực tế sau khi loại bỏ omit
        valid_indices = [
            i for i in range(plan.img_start, plan.img_end + 1)
            if i not in plan.img_omit
        ]

        # Chạy OCR cho thiết bị này
        ocr_vals, ocr_warns = read_device_ocr(
            bmp_indices=valid_indices,
            bmp_map=bmp_map,
        )
        if ocr_warns:
            warnings_out.extend([f"[{plan.device_raw}] {w}" for w in ocr_warns])

        # Ghi từng giá trị vào Excel
        for field_name, value in ocr_vals.items():
            if value is None or field_name not in col_indices:
                continue
            
            col_idx = col_indices[field_name]
            cell = ws.cell(row=plan.excel_row, column=col_idx)

            # Không ghi đè nếu ô đã có giá trị (trừ khi overwrite_existing)
            if not overwrite_existing and cell.value is not None and str(cell.value).strip():
                continue

            cell.value = round(value, 4)

    try:
        wb.save(excel_path)
    except Exception as e:
        warnings_out.append(f"OCR: không lưu được Excel sau khi cập nhật ({e}).")

    return warnings_out


def process_field_zip_bytes(
    zip_bytes: bytes,
    work_dir: str,
    run_ocr: bool = True,
    ocr_overwrite: bool = False,
) -> tuple[str, list[str], list[str]]:
    """Xử lý toàn bộ quy trình tổ chức hồ sơ từ dữ liệu ZIP.

    Giải nén, đọc kế hoạch, sắp xếp thư mục, chạy OCR và nén lại kết quả.

    Args:
        zip_bytes (bytes): Dữ liệu nhị phân của file ZIP đầu vào.
        work_dir (str): Thư mục làm việc tạm thời.
        run_ocr (bool): Có chạy OCR hay không.
        ocr_overwrite (bool): Có ghi đè dữ liệu OCR hay không.

    Returns:
        tuple[str, list[str], list[str]]: Một tuple gồm:
            - Đường dẫn file ZIP kết quả.
            - Danh sách các cảnh báo.
            - Danh sách các lỗi nghiêm trọng (khiến quy trình dừng lại).
    """
    warnings: list[str] = []
    extract = os.path.join(work_dir, "in")
    os.makedirs(extract, exist_ok=True)
    bio = io.BytesIO(zip_bytes)
    with zipfile.ZipFile(bio, "r") as zf:
        zf.extractall(extract)

    excel_path = find_first_excel(extract)
    if not excel_path:
        return "", warnings, ["Không tìm thấy file Excel (.xlsx/.xlsm/.xls) trong ZIP."]

    s_map, s_err = scan_s_folders(extract)
    warnings.extend(s_err)
    bmp_map, bmp_dup = scan_bmp_files(extract)
    warnings.extend(bmp_dup)

    try:
        plans, w2 = read_plans_from_excel(excel_path)
        warnings.extend(w2)
    except ValueError as e:
        return "", warnings, [str(e)]

    fatal = validate_plans_against_fs(plans, s_map, bmp_map)
    if fatal:
        return "", warnings, fatal

    # ── Chạy OCR tự động điền Excel ──────────────────────────────────────────
    if run_ocr:
        ocr_warns = run_ocr_and_update_excel(
            excel_path=excel_path,
            plans=plans,
            bmp_map=bmp_map,
            overwrite_existing=ocr_overwrite,
        )
        warnings.extend(ocr_warns)

    staging = os.path.join(work_dir, "staging")
    os.makedirs(staging, exist_ok=True)
    build_project_output(extract, staging, plans, s_map, bmp_map, warnings)

    # Copy Excel (đã cập nhật nếu run_ocr=True) vào Project_Output để có trong ZIP kết quả
    excel_dst = os.path.join(staging, "Project_Output", os.path.basename(excel_path))
    if not os.path.exists(excel_dst):
        shutil.copy2(excel_path, excel_dst)

    out_zip = os.path.join(work_dir, "KEW_HoSoDaXuLy.zip")
    proj = os.path.join(staging, "Project_Output")
    zip_directory(proj, out_zip)
    return out_zip, warnings, []
