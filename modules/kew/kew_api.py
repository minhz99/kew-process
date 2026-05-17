"""
Module: kew_api.py
Description: Cung cấp các API RESTful để xử lý dữ liệu từ thiết bị KEW6315.
Bao gồm các tính năng: xuất dữ liệu sang template MBA.xlsm và tổ chức hồ sơ hiện trường từ ZIP.
"""

import os
import io
import shutil
import tempfile
import traceback
import zipfile
import re
import urllib.parse
from pathlib import Path
from flask import Blueprint, request, jsonify, send_file


try:
    import pandas as pd
    from openpyxl import load_workbook
    from typing import Mapping
    _MBA_DEPS_OK = True
except ImportError:
    _MBA_DEPS_OK = False

# ─── Cấu hình MBA export ─────────────────────────────────────────────────────
from modules.report.gen_excel_mba import (
    _MBA_SKIP_ROWS,
    _mba_extract,
    _mba_write
)

_MBA_TEMPLATE_PATH = os.path.join(
    os.path.dirname(os.path.abspath(__file__)),
    '..', '..', 'static', 'excel-template', 'MBA.xlsm'
)
_MBA_PREBUILT_COUNT = 10  # Số sheet có sẵn trong template (MBA1 … MBA10)

kew_bp = Blueprint('kew_bp', __name__)


@kew_bp.route('/export-mba', methods=['POST'])
def export_mba():
    """
    API endpoint để xuất báo cáo MBA sang file Excel (.xlsm).
    
    Nhận các file KEW hoặc ZIP từ người dùng, trích xuất dữ liệu và điền vào 
    template Excel có sẵn. Mỗi file KEW sẽ tương ứng với một sheet MBA trong file kết quả.
    
    Returns:
        Response: File Excel kết quả hoặc lỗi JSON.
    """
    if not _MBA_DEPS_OK:
        return jsonify({'error': 'Thiếu thư viện pandas hoặc openpyxl.'}), 500

    out_filename = request.form.get('filename', '').strip() or 'NX-MBA.xlsm'
    # Đảm bảo đuôi .xlsm để bảo toàn macro của template
    if not out_filename.lower().endswith(('.xlsx', '.xlsm')):
        out_filename += '.xlsm'

    # ── Lấy tất cả file KEW bytes từ request ─────────────────────────────────
    def _extract_kew_bytes(f) -> list:
        """Từ 1 FileStorage trả về list [(name, bytes)]."""
        raw = f.read()
        if f.filename.upper().endswith('.ZIP'):
            with zipfile.ZipFile(io.BytesIO(raw), 'r', metadata_encoding="utf-8") as zf:
                entries = [n for n in zf.namelist() if n.upper().endswith('.KEW')]
                return [(os.path.basename(n), zf.read(n)) for n in entries]
        return [(os.path.basename(f.filename), raw)]

    kew_list = []  # list of (name, bytes)
    if 'files' in request.files:
        for f in request.files.getlist('files'):
            kew_list.extend(_extract_kew_bytes(f))
    elif 'zip' in request.files:
        kew_list.extend(_extract_kew_bytes(request.files['zip']))

    if not kew_list:
        return jsonify({'error': 'Cần upload ít nhất 1 file .KEW hoặc .ZIP.'}), 400


    # ── Load template ─────────────────────────────────────────────────────────
    template_path = os.path.normpath(_MBA_TEMPLATE_PATH)
    if not os.path.isfile(template_path):
        return jsonify({'error': f'Không tìm thấy template MBA.xlsm tại {template_path}'}), 500

    try:
        wb = load_workbook(template_path, keep_vba=True)
    except Exception as e:
        return jsonify({'error': f'Không mở được template: {e}'}), 500

    # ── Danh sách sheet có sẵn trong template (MBA1 … MBA10) ─────────────────
    # Template đã có sẵn _MBA_PREBUILT_COUNT sheet; chỉ cần trỏ vào và đổi tên.
    # Nếu số lượng MBA vượt quá số sheet có sẵn → copy_worksheet từ sheet cuối.
    prebuilt_sheets = wb.sheetnames[:_MBA_PREBUILT_COUNT]

    errors_list = []
    try:
        for idx, (kew_name, kew_bytes) in enumerate(kew_list):
            # ── Parse KEW ────────────────────────────────────────────────────────
            try:
                df_raw = pd.read_csv(
                    io.BytesIO(kew_bytes),
                    skiprows=_MBA_SKIP_ROWS,
                    low_memory=False,
                )
                df_raw.columns = df_raw.columns.str.strip()
            except Exception as e:
                errors_list.append(f'{kew_name}: không đọc được ({e})')
                continue

            df, warnings = _mba_extract(df_raw)
            if warnings:
                errors_list.extend([f"{kew_name}: {w}" for w in warnings])

            # ── Lấy sheet đích ───────────────────────────────────────────────────
            try:
                if idx < len(prebuilt_sheets):
                    # Tái sử dụng sheet có sẵn, KHÔNG đổi tên
                    ws = wb[prebuilt_sheets[idx]]
                else:
                    # Vượt quá số sheet template → copy từ sheet cuối cùng có sẵn
                    ws_ref = wb[wb.sheetnames[len(prebuilt_sheets) - 1]]
                    ws = wb.copy_worksheet(ws_ref)
                    ws.title = f'MBA{idx + 1}'
            except Exception as e:
                errors_list.append(f'{kew_name}: không thể lấy sheet ({e})')
                continue

            try:
                _mba_write(ws, df)
            except Exception as e:
                errors_list.append(f'{kew_name}: lỗi ghi dữ liệu ({e})')
                continue

        if errors_list and len(errors_list) == len(kew_list):
            # Tất cả đều lỗi
            return jsonify({'error': 'Tất cả file thất bại: ' + '; '.join(errors_list)}), 400

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': f'Lỗi hệ thống khi xử lý: {str(e)}'}), 500

    # Giữ đuôi .xlsm để bảo toàn macro
    if not out_filename.lower().endswith('.xlsm'):
        out_filename = os.path.splitext(out_filename)[0] + '.xlsm'

    resp = send_file(
        output,
        as_attachment=True,
        download_name=out_filename,
        mimetype='application/vnd.ms-excel.sheet.macroEnabled.12',
    )
    if errors_list:
        resp.headers['X-MBA-Warnings'] = urllib.parse.quote('; '.join(errors_list))
    return resp


@kew_bp.route("/organize-field-zip", methods=["POST"])
def organize_field_zip():
    """
    API endpoint để tổ chức lại file ZIP hồ sơ hiện trường.
    
    Thực hiện:
    1. Đọc file Excel kế hoạch trong ZIP.
    2. Đổi tên các thư mục Sxxxx thành tên thiết bị tương ứng.
    3. Di chuyển các file ảnh PS-SDxxx.BMP vào đúng thư mục thiết bị.
    4. (Tùy chọn) Chạy OCR tự động đọc các thông số đo từ ảnh BMP và ghi vào Excel.
    5. Nén lại thành file ZIP kết quả.

    Form params:
        zip / file: File ZIP đầu vào.
        run_ocr: '1' hoặc 'true' để chạy OCR sau khi sắp xếp (mặc định: true).
        ocr_overwrite: '1' hoặc 'true' để ghi đè các ô Excel đã có giá trị (mặc định: false).
    
    Returns:
        Response: File ZIP đã tổ chức lại hoặc lỗi JSON.
    """
    from modules.kew import organize_field_zip as organize_mod

    zf = request.files.get("zip") or request.files.get("file")
    if zf is None or not getattr(zf, "filename", None):
        return jsonify({"error": "Cần upload file ZIP (form field zip hoặc file)."}), 400
    if not str(zf.filename).lower().endswith(".zip"):
        return jsonify({"error": "Chỉ chấp nhận file .zip."}), 400

    zip_bytes = zf.read()
    if not zip_bytes:
        return jsonify({"error": "File ZIP rỗng."}), 400

    # Tham số OCR từ form
    run_ocr_raw = request.form.get("run_ocr", "true").strip().lower()
    run_ocr = run_ocr_raw not in ("0", "false", "no", "")
    ocr_overwrite_raw = request.form.get("ocr_overwrite", "false").strip().lower()
    ocr_overwrite = ocr_overwrite_raw in ("1", "true", "yes")

    work = tempfile.mkdtemp(prefix="kew_field_org_")
    try:
        out_path, warnings, fatal = organize_mod.process_field_zip_bytes(
            zip_bytes, work, run_ocr=run_ocr, ocr_overwrite=ocr_overwrite
        )
        if fatal:
            return jsonify({"errors": fatal, "warnings": warnings}), 400
        if not out_path or not os.path.isfile(out_path):
            return jsonify({"error": "Không tạo được file kết quả.", "warnings": warnings}), 500
        with open(out_path, "rb") as fh:
            buf = io.BytesIO(fh.read())
        buf.seek(0)
        resp = send_file(
            buf,
            as_attachment=True,
            download_name="KEW_HoSoDaXuLy.zip",
            mimetype="application/zip",
        )
        if warnings:
            resp.headers["X-KEW-Field-Warnings"] = urllib.parse.quote("; ".join(warnings))
        return resp
    except Exception as e:
        traceback.print_exc()
        return jsonify({"error": f"Lỗi xử lý hồ sơ KEW: {e}"}), 500
    finally:
        shutil.rmtree(work, ignore_errors=True)


@kew_bp.route("/generate-word-report", methods=["POST"])
def generate_word_report():
    """
    API endpoint để sinh báo cáo Word tổng hợp từ hồ sơ thiết bị (file ZIP).
    
    Quy trình xử lý:
    1. Tiếp nhận file ZIP từ client và giải nén vào bộ nhớ tạm.
    2. Duyệt cấu trúc thư mục, nhận diện Máy biến áp (MBA) và thiết bị phụ tải.
    3. Tạo và trộn các section theo thứ tự: Các MBA -> Bảng tổng kết MBA -> Các thiết bị phụ tải.
    4. Trả về file Word hoàn chỉnh cho người dùng.
    
    Returns:
        Response: File .docx báo cáo (application/vnd.openxmlformats-officedocument.wordprocessingml.document) hoặc lỗi JSON.
    """
    from modules.report.gen_word import build_word_report_from_zip

    zf = request.files.get("zip") or request.files.get("file")
    if zf is None or not getattr(zf, "filename", None):
        return jsonify({"error": "Cần upload file ZIP (form field zip hoặc file)."}), 400
    if not str(zf.filename).lower().endswith(".zip"):
        return jsonify({"error": "Chỉ chấp nhận file .zip."}), 400

    zip_bytes = zf.read()
    if not zip_bytes:
        return jsonify({"error": "File ZIP rỗng."}), 400

    out_name = (request.form.get("filename", "") or "").strip() or "BaoCao_KEW.docx"
    if not out_name.lower().endswith(".docx"):
        out_name += ".docx"

    work = tempfile.mkdtemp(prefix="kew_word_")
    try:
        out_path = os.path.join(work, out_name)
        try:
            _, warnings = build_word_report_from_zip(zip_bytes, out_path)
        except (FileNotFoundError, ValueError) as e:
            return jsonify({"error": str(e)}), 400
        except RuntimeError as e:
            return jsonify({"error": str(e)}), 400
        if not os.path.isfile(out_path):
            return jsonify({"error": "Không tạo được file Word."}), 500

        with open(out_path, "rb") as fh:
            buf = io.BytesIO(fh.read())
        buf.seek(0)
        resp = send_file(
            buf,
            as_attachment=True,
            download_name=out_name,
            mimetype="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
        )
        if warnings:
            resp.headers["X-KEW-Word-Warnings"] = urllib.parse.quote("; ".join(warnings))
        return resp
    except Exception as e:
        traceback.print_exc()
        return jsonify({"error": f"Lỗi sinh báo cáo Word: {e}"}), 500
    finally:
        shutil.rmtree(work, ignore_errors=True)


@kew_bp.route("/generate-chapter4", methods=["POST"])
def generate_chapter4():
    """
    API endpoint để sinh Chương 4 Word: chỉ các thiết bị có ``type=\"4\"``.

    Sử dụng template ``device4.docx`` (cấu trúc giống ``device.docx``).

    Returns:
        Response: File .docx Chương 4 hoặc lỗi JSON.
    """
    from modules.report.gen_word import build_chapter4_from_zip

    zf = request.files.get("zip") or request.files.get("file")
    if zf is None or not getattr(zf, "filename", None):
        return jsonify({"error": "Cần upload file ZIP (form field zip hoặc file)."}), 400
    if not str(zf.filename).lower().endswith(".zip"):
        return jsonify({"error": "Chỉ chấp nhận file .zip."}), 400

    zip_bytes = zf.read()
    if not zip_bytes:
        return jsonify({"error": "File ZIP rỗng."}), 400

    out_name = (request.form.get("filename", "") or "").strip() or "Chương 4.docx"
    if not out_name.lower().endswith(".docx"):
        out_name += ".docx"

    work = tempfile.mkdtemp(prefix="kew_chap4_")
    try:
        out_path = os.path.join(work, out_name)
        try:
            _, warnings = build_chapter4_from_zip(zip_bytes, out_path)
        except (FileNotFoundError, ValueError) as e:
            return jsonify({"error": str(e)}), 400
        except RuntimeError as e:
            return jsonify({"error": str(e)}), 400
        if not os.path.isfile(out_path):
            return jsonify({"error": "Không tạo được file Word Chương 4."}), 500

        with open(out_path, "rb") as fh:
            buf = io.BytesIO(fh.read())
        buf.seek(0)
        resp = send_file(
            buf,
            as_attachment=True,
            download_name=out_name,
            mimetype="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
        )
        if warnings:
            resp.headers["X-KEW-Word-Warnings"] = urllib.parse.quote("; ".join(warnings))
        return resp
    except Exception as e:
        traceback.print_exc()
        return jsonify({"error": f"Lỗi sinh báo cáo Chương 4: {e}"}), 500
    finally:
        shutil.rmtree(work, ignore_errors=True)


@kew_bp.route("/generate-chapter5", methods=["POST"])
def generate_chapter5():
    """
    API endpoint để sinh Chương 5 Word: MBA + các thiết bị không có ``type=\"4\"``.

    Returns:
        Response: File .docx Chương 5 hoặc lỗi JSON.
    """
    from modules.report.gen_word import build_chapter5_from_zip

    zf = request.files.get("zip") or request.files.get("file")
    if zf is None or not getattr(zf, "filename", None):
        return jsonify({"error": "Cần upload file ZIP (form field zip hoặc file)."}), 400
    if not str(zf.filename).lower().endswith(".zip"):
        return jsonify({"error": "Chỉ chấp nhận file .zip."}), 400

    zip_bytes = zf.read()
    if not zip_bytes:
        return jsonify({"error": "File ZIP rỗng."}), 400

    out_name = (request.form.get("filename", "") or "").strip() or "Chương 5.docx"
    if not out_name.lower().endswith(".docx"):
        out_name += ".docx"

    work = tempfile.mkdtemp(prefix="kew_chap5_")
    try:
        out_path = os.path.join(work, out_name)
        try:
            _, warnings = build_chapter5_from_zip(zip_bytes, out_path)
        except (FileNotFoundError, ValueError) as e:
            return jsonify({"error": str(e)}), 400
        except RuntimeError as e:
            return jsonify({"error": str(e)}), 400
        if not os.path.isfile(out_path):
            return jsonify({"error": "Không tạo được file Word Chương 5."}), 500

        with open(out_path, "rb") as fh:
            buf = io.BytesIO(fh.read())
        buf.seek(0)
        resp = send_file(
            buf,
            as_attachment=True,
            download_name=out_name,
            mimetype="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
        )
        if warnings:
            resp.headers["X-KEW-Word-Warnings"] = urllib.parse.quote("; ".join(warnings))
        return resp
    except Exception as e:
        traceback.print_exc()
        return jsonify({"error": f"Lỗi sinh báo cáo Chương 5: {e}"}), 500
    finally:
        shutil.rmtree(work, ignore_errors=True)


@kew_bp.route("/generate-table6", methods=["POST"])
def generate_table6():
    """
    API endpoint để sinh bảng tổng hợp kết quả đo kiểm (Table 6).
    
    Trích xuất các chỉ số I, P, PF, THD, TDD từ Excel hiện trường và 
    tạo bảng tổng hợp trong file Word.
    
    Returns:
        Response: File Word chứa Table 6 hoặc lỗi JSON.
    """
    from modules.report.gen_word import generate_table6_from_zip

    zf = request.files.get("zip") or request.files.get("file")
    if zf is None or not getattr(zf, "filename", None):
        return jsonify({"error": "Cần upload file ZIP (form field zip hoặc file)."}), 400
    if not str(zf.filename).lower().endswith(".zip"):
        return jsonify({"error": "Chỉ chấp nhận file .zip."}), 400

    zip_bytes = zf.read()
    if not zip_bytes:
        return jsonify({"error": "File ZIP rỗng."}), 400

    out_name = (request.form.get("filename", "") or "").strip() or "Bang_TongHop_Table6.docx"
    if not out_name.lower().endswith(".docx"):
        out_name += ".docx"
    work = tempfile.mkdtemp(prefix="kew_table6_")
    try:
        out_path = os.path.join(work, out_name)
        try:
            _, warnings = generate_table6_from_zip(zip_bytes, out_path)
        except (FileNotFoundError, ValueError) as e:
            return jsonify({"error": str(e)}), 400
        except RuntimeError as e:
            return jsonify({"error": str(e)}), 400
        if not os.path.isfile(out_path):
            return jsonify({"error": "Không tạo được bảng tổng hợp."}), 500

        with open(out_path, "rb") as fh:
            buf = io.BytesIO(fh.read())
        buf.seek(0)
        resp = send_file(
            buf,
            as_attachment=True,
            download_name=out_name,
            mimetype="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
        )
        if warnings:
            resp.headers["X-KEW-Table6-Warnings"] = urllib.parse.quote("; ".join(warnings))
        return resp
    except Exception as e:
        traceback.print_exc()
        return jsonify({"error": f"Lỗi sinh bảng tổng hợp: {e}"}), 500
    finally:
        shutil.rmtree(work, ignore_errors=True)
@kew_bp.route("/generate-excel-mba", methods=["POST"])
def generate_excel_mba():
    """
    API endpoint để sinh báo cáo Excel MBA từ hồ sơ thiết bị (file ZIP).
    Sử dụng template MBA.xlsm với logic copy sheet và bảng tổng hợp.
    """
    from modules.report.gen_word import _find_project_root, _find_first_excel, read_device_metadata_from_excel, _lookup_device_metadata, _nfc
    from modules.report.gen_excel_mba import generate_mba_excel_from_devices

    zf = request.files.get("zip") or request.files.get("file")
    if zf is None or not getattr(zf, "filename", None):
        return jsonify({"error": "Cần upload file ZIP (form field zip hoặc file)."}), 400
    
    zip_bytes = zf.read()
    if not zip_bytes:
        return jsonify({"error": "File ZIP rỗng."}), 400

    out_name = (request.form.get("filename", "") or "").strip() or "BaoCao_MBA.xlsm"
    if not out_name.lower().endswith((".xlsm", ".xlsx")):
        out_name += ".xlsm"

    work = tempfile.mkdtemp(prefix="kew_excel_mba_")
    try:
        extract = os.path.join(work, "in")
        os.makedirs(extract, exist_ok=True)
        bio = io.BytesIO(zip_bytes)
        with zipfile.ZipFile(bio, "r", metadata_encoding="utf-8") as zf_in:
            zf_in.extractall(extract)

        project_root = _find_project_root(Path(extract))
        excel_path = _find_first_excel(Path(extract))
        metadata = read_device_metadata_from_excel(excel_path) if excel_path else {}

        raw_dirs = [
            d for d in project_root.iterdir()
            if d.is_dir() and not d.name.startswith(".") and d.name != "__MACOSX"
        ]

        _stt_fallback = 10**9

        def _device_dir_sort_key(p: Path) -> tuple[int, int, str]:
            from modules.report.gen_word import _resolve_word_section_kind
            m = _lookup_device_metadata(metadata, p.name)
            display = _nfc(m.get("name") or p.name)
            kind = _resolve_word_section_kind({"kind": m.get("kind")}, name=display, default_kind=None)
            st = m.get("stt")
            st_val = st if isinstance(st, int) else _stt_fallback
            return (0 if kind == "mba" else 1, st_val, p.name.lower())

        device_dirs = sorted(raw_dirs, key=_device_dir_sort_key)
        
        devices = []
        for d in device_dirs:
            meta = _lookup_device_metadata(metadata, d.name)
            display = _nfc(meta.get("name") or d.name)
            devices.append({
                "name": display,
                "folder": d,
                "kind": meta.get("kind"),
                "excel_params": meta.get("excel_params") or {},
            })

        template_path = _MBA_TEMPLATE_PATH
        out_path = os.path.join(work, out_name)
        
        try:
            generate_mba_excel_from_devices(devices, out_path, template_path)
        except Exception as e:
            traceback.print_exc()
            return jsonify({"error": f"Lỗi sinh báo cáo Excel MBA: {e}"}), 500

        if not os.path.isfile(out_path):
            return jsonify({"error": "Không tạo được file Excel MBA."}), 500

        with open(out_path, "rb") as fh:
            buf = io.BytesIO(fh.read())
        buf.seek(0)
        
        return send_file(
            buf,
            as_attachment=True,
            download_name=out_name,
            mimetype="application/vnd.ms-excel.sheet.macroEnabled.12" if out_name.endswith(".xlsm") else "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    except Exception as e:
        traceback.print_exc()
        return jsonify({"error": f"Lỗi hệ thống: {e}"}), 500
    finally:
        shutil.rmtree(work, ignore_errors=True)
