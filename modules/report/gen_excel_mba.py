"""
Module: gen_excel_mba.py
Description: Chứa các hàm hỗ trợ sinh báo cáo Excel cho Máy biến áp (MBA).
Hỗ trợ đọc dữ liệu từ file INPSxxxx.KEW và điền vào template Excel MBA.
"""
import os
import io
import unicodedata
import re
from pathlib import Path
from typing import Sequence, Mapping
from openpyxl import load_workbook

def _nfc(s: object) -> str:
    """Chuẩn hóa tên hiển thị về dạng NFC.

    Args:
        s (object): Đối tượng cần chuẩn hóa (thường là chuỗi).

    Returns:
        str: Chuỗi đã được chuẩn hóa.
    """
    if s is None:
        return ""
    return unicodedata.normalize("NFC", str(s).strip())

def _parse_float(v: object) -> float | None:
    """Trích xuất giá trị số thực từ ô dữ liệu.

    Hỗ trợ chuyển đổi từ chuỗi, xử lý dấu phẩy thành dấu chấm.

    Args:
        v (object): Giá trị cần trích xuất.

    Returns:
        Optional[float]: Giá trị số thực hoặc None nếu không hợp lệ.
    """
    if v is None:
        return None
    try:
        import pandas as pd
        if isinstance(v, float) and pd.isna(v):
            return None
    except Exception:
        pass
    try:
        s = str(v).strip().replace(",", ".")
        if not s:
            return None
        return float(s)
    except (ValueError, TypeError):
        return None

_MBA_SKIP_ROWS = 1
_MBA_START_ROW = 2
_MBA_START_COL = 2

_MBA_COLUMN_MAPPING = {
    "AVG_A1[A]": "AVG_A1[A]",
    "AVG_A2[A]": "AVG_A2[A]",
    "AVG_A3[A]": "AVG_A3[A]",
    "AVG_P[W]": "AVG_P[W]",
    "AVG_Q[var]": "AVG_Q[var]",
    "AVG_S[VA]": "AVG_S[VA]",
    "AVG_PF[_]": "AVG_PF",
    "AVG_VL1[V]": "AVG_VL1[V]",
    "AVG_VL2[V]": "AVG_VL2[V]",
    "AVG_VL3[V]": "AVG_VL3[V]",
    "AVG_THDVR1[%]": "AVG_Vthd1[%]",
    "AVG_THDVR2[%]": "AVG_Vthd2[%]",
    "AVG_THDVR3[%]": "AVG_Vthd3[%]",
    "AVG_THDAR1[%]": "AVG_Athd1[%]",
    "AVG_THDAR2[%]": "AVG_Athd2[%]",
    "AVG_THDAR3[%]": "AVG_Athd3[%]",
    "AVG_UV[%]": "AVG_Vunb[%]",
    "AVG_UA[%]": "AVG_Aunb[%]",
}
_MBA_SCALE_DIV_1000 = ("AVG_P[W]", "AVG_Q[var]", "AVG_S[VA]")
_MBA_ROUND = {
    "AVG_A1[A]": 2, "AVG_A2[A]": 2, "AVG_A3[A]": 2,
    "AVG_P[W]": 2, "AVG_Q[var]": 2, "AVG_S[VA]": 2,
    "AVG_PF": 4,
    "AVG_VL1[V]": 1, "AVG_VL2[V]": 1, "AVG_VL3[V]": 1,
    "AVG_Vthd1[%]": 3, "AVG_Vthd2[%]": 3, "AVG_Vthd3[%]": 3,
    "AVG_Athd1[%]": 2, "AVG_Athd2[%]": 2, "AVG_Athd3[%]": 2,
    "AVG_Vunb[%]": 4, "AVG_Aunb[%]": 3,
}
_MBA_FMT = {
    "AVG_A1[A]": "0.00", "AVG_A2[A]": "0.00", "AVG_A3[A]": "0.00",
    "AVG_P[W]": "0.00", "AVG_Q[var]": "0.00", "AVG_S[VA]": "0.00",
    "AVG_PF": "0.0000",
    "AVG_VL1[V]": "0.0", "AVG_VL2[V]": "0.0", "AVG_VL3[V]": "0.0",
    "AVG_Vthd1[%]": "0.000", "AVG_Vthd2[%]": "0.000", "AVG_Vthd3[%]": "0.000",
    "AVG_Athd1[%]": "0.00", "AVG_Athd2[%]": "0.00", "AVG_Athd3[%]": "0.00",
    "AVG_Vunb[%]": "0.0000", "AVG_Aunb[%]": "0.000",
}

def _mba_to_number(v):
    """Chuyển đổi một giá trị từ file KEW sang số thực (float).
    
    Xử lý các đơn vị k (kilo), m (mega) và loại bỏ các ký hiệu đơn vị khác.
    
    Args:
        v: Giá trị cần chuyển đổi.
        
    Returns:
        float or pd.NA: Giá trị số sau khi chuyển đổi hoặc pd.NA nếu không hợp lệ.
    """
    import pandas as pd
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return pd.NA
    s = str(v).strip()
    s = s.replace('-', '')
    if not s or s.lower() == "nan":
        return pd.NA
    m = re.search(r"(\d+(?:\.\d+)?)\s*([kKmM])?", s)
    if m:
        base = float(m.group(1))
        unit = (m.group(2) or "").lower()
        if unit == "k":
            base *= 1_000.0
        elif unit == "m":
            base *= 1_000_000.0
        return base
    return pd.to_numeric(s, errors="coerce")

def _mba_extract(df: "pd.DataFrame") -> "tuple[pd.DataFrame, list[str]]":
    """Trích xuất và chuẩn hóa dữ liệu từ DataFrame gốc của file KEW.
    
    Áp dụng mapping cột, chuyển đổi đơn vị và làm tròn số.
    
    Args:
        df (pd.DataFrame): DataFrame dữ liệu gốc.
        
    Returns:
        tuple: (DataFrame đã chuẩn hóa, danh sách các cảnh báo warnings).
    """
    import pandas as pd
    orig_cols = list(_MBA_COLUMN_MAPPING.keys())
    available_cols = [c for c in orig_cols if c in df.columns]
    missing = [c for c in orig_cols if c not in df.columns]
    
    out = df[available_cols].copy()
    rename_mapping = {k: v for k, v in _MBA_COLUMN_MAPPING.items() if k in available_cols}
    out = out.rename(columns=rename_mapping)
    
    for m in missing:
        out[_MBA_COLUMN_MAPPING[m]] = pd.NA

    target_ordered = list(_MBA_COLUMN_MAPPING.values())
    out = out[target_ordered]

    for col in out.columns:
        out[col] = out[col].map(_mba_to_number)
        
    cols_to_check = [
        "AVG_A1[A]", "AVG_A2[A]", "AVG_A3[A]",
        "AVG_P[W]", "AVG_Q[var]", "AVG_S[VA]"
    ]
    for col in cols_to_check:
        renamed = dict(_MBA_COLUMN_MAPPING).get(col, col)
        if renamed in out.columns:
            out[renamed] = out[renamed].apply(lambda x: x * 1000.0 if pd.notna(x) and abs(x) < 10.0 else x)

    for col in _MBA_SCALE_DIV_1000:
        renamed = dict(_MBA_COLUMN_MAPPING).get(col, col)
        if renamed in out.columns:
            out[renamed] = out[renamed] / 1000.0
    for col, nd in _MBA_ROUND.items():
        if col in out.columns:
            out[col] = pd.to_numeric(out[col], errors='coerce').round(nd)
            
    warnings = []
    if missing:
        warnings.append(f"Dữ liệu gốc khuyết {len(missing)} cột: {missing}")
        
    return out, warnings

def _evaluate_for_excel(df: "pd.DataFrame") -> dict[int, str]:
    """Đánh giá các chỉ số kỹ thuật (điện áp, PF, THD, unbalance) để ghi nhận xét vào file Excel.
    
    Sử dụng các hàm đánh giá dùng chung với module report.
    
    Args:
        df (pd.DataFrame): DataFrame dữ liệu đã chuẩn hóa.
        
    Returns:
        dict: Một dictionary mapping dòng (row index) với chuỗi nhận xét (ví dụ {8: "Đạt"}).
    """
    from modules.report.gen_word import (
        _eval_voltage, _eval_pf, _eval_thd, _eval_unbalance,
        _V_DEV_LIMIT_PCT, _THDV_LIMIT_PCT, _TDD_LIMIT_PCT, _MBA_NOMINAL_VOLTAGE_V
    )
    
    res = {}
    
    if "AVG_VL1[V]" in df.columns:
        u_vals = df["AVG_VL1[V]"].dropna()
        if not u_vals.empty:
            eval_str, _, _, _ = _eval_voltage(u_vals.max(), u_vals.min(), u_vals.mean(), _MBA_NOMINAL_VOLTAGE_V)
            res[8] = eval_str
            
    if "AVG_Vunb[%]" in df.columns:
        vu_vals = df["AVG_Vunb[%]"].dropna()
        if not vu_vals.empty:
            res[14] = _eval_unbalance(vu_vals.max(), vu_vals.mean(), _V_DEV_LIMIT_PCT)
            
    if "AVG_PF" in df.columns:
        pf_vals = df["AVG_PF"].dropna()
        if not pf_vals.empty:
            res[16] = _eval_pf(pf_vals.max(), pf_vals.min(), pf_vals.mean())
            
    thd_cols = [c for c in ["AVG_Vthd1[%]", "AVG_Vthd2[%]", "AVG_Vthd3[%]"] if c in df.columns]
    if thd_cols:
        max_vals = [df[c].max() for c in thd_cols if not df[c].dropna().empty]
        avg_vals = [df[c].mean() for c in thd_cols if not df[c].dropna().empty]
        res[20] = _eval_thd(max_vals, avg_vals, _THDV_LIMIT_PCT)
        
    tdd_cols = [c for c in ["AVG_Athd1[%]", "AVG_Athd2[%]", "AVG_Athd3[%]"] if c in df.columns]
    if tdd_cols:
        max_vals = [df[c].max() for c in tdd_cols if not df[c].dropna().empty]
        avg_vals = [df[c].mean() for c in tdd_cols if not df[c].dropna().empty]
        res[23] = _eval_thd(max_vals, avg_vals, _TDD_LIMIT_PCT)
        
    return res

def _mba_write(ws, df: "pd.DataFrame") -> None:
    """Ghi dữ liệu từ DataFrame vào worksheet Excel và áp dụng định dạng.
    
    Cũng ghi các nhận xét đánh giá vào cột AE (cột 31).
    
    Args:
        ws: Worksheet của openpyxl.
        df (pd.DataFrame): DataFrame dữ liệu đã chuẩn hóa.
    """
    import pandas as pd
    sr, sc = _MBA_START_ROW, _MBA_START_COL
    
    last_row = ws.max_row
    last_col = sc + len(df.columns) - 1
    for r in range(sr + 1, last_row + 1):
        for c in range(sc, last_col + 1):
            cell = ws.cell(row=r, column=c)
            cell.value = None
            cell.number_format = "General"
            
    for ci, cname in enumerate(df.columns):
        ws.cell(row=sr, column=sc + ci, value=cname)
    for ri, row in enumerate(df.values):
        for ci, val in enumerate(row):
            if pd.notna(val):
                cname = df.columns[ci]
                cell = ws.cell(row=sr + 1 + ri, column=sc + ci, value=float(val))
                fmt = _MBA_FMT.get(cname)
                if fmt:
                    cell.number_format = fmt

    evals = _evaluate_for_excel(df)
    for r_idx, ev_str in evals.items():
        cell = ws.cell(row=r_idx, column=31)
        cell.value = ev_str

def generate_mba_excel_from_devices(
    devices: Sequence[Mapping],
    output_path: str | Path,
    template_path: str | Path,
) -> Path:
    """Sinh báo cáo Excel MBA từ danh sách metadata thiết bị.

    Đọc dữ liệu từ metadata, điền vào các sheet tương ứng trong template.

    Args:
        devices (Sequence[Mapping]): Danh sách metadata của các thiết bị.
        output_path (str | Path): Đường dẫn lưu file Excel kết quả.
        template_path (str | Path): Đường dẫn đến file template Excel MBA.

    Returns:
        Path: Đường dẫn đến file Excel kết quả đã tạo.

    Raises:
        ValueError: Nếu không tìm thấy máy biến áp nào hoặc template thiếu sheet 'MBA1'.
    """
    from modules.report.gen_word import (
        _resolve_word_section_kind, _eval_voltage, _eval_pf, _eval_thd, _eval_unbalance,
        _V_DEV_LIMIT_PCT, _THDV_LIMIT_PCT, _TDD_LIMIT_PCT, _MBA_NOMINAL_VOLTAGE_V
    )

    # 1. Lọc danh sách MBA
    mbas = []
    found_kinds = []
    for d in devices:
        k_val = d.get("kind")
        spec = {"kind": k_val} if k_val is not None else {}
        kind = _resolve_word_section_kind(spec, name=d.get("name", ""), default_kind=None)
        found_kinds.append(f"{d.get('name')}: {kind} (original type: {k_val})")
        if kind == "mba":
            mbas.append(d)

    if not mbas:
        details = "; ".join(found_kinds)
        raise ValueError(f"Không tìm thấy máy biến áp nào trong dữ liệu. Danh sách thiết bị: {details}")

    # 2. Mở template
    wb = load_workbook(str(template_path), keep_vba=True)
    
    # Đảm bảo sheet "MBA1" tồn tại
    if "MBA1" not in wb.sheetnames:
        # Nếu template không có MBA1, cố gắng lấy sheet đầu tiên hoặc tạo mới (nhưng spec yêu cầu MBA1)
        raise ValueError("Template Excel không có sheet 'MBA1'.")

    # Xóa các sheet MBA khác (nếu có sẵn trong template cũ) để làm sạch
    for sname in list(wb.sheetnames):
        if sname.startswith("MBA") and sname != "MBA1":
            del wb[sname]

    source_ws = wb["MBA1"]
    
    # 3. Điền dữ liệu vào các sheet MBA
    for i, mba_data in enumerate(mbas):
        name = _nfc(mba_data.get("name") or f"MBA{i+1}")
        # Hạn chế độ dài tên sheet (Excel giới hạn 31 ký tự)
        sheet_name = name[:31]
        
        if i == 0:
            ws = source_ws
            ws.title = sheet_name
        else:
            ws = wb.copy_worksheet(source_ws)
            ws.title = sheet_name

        ep = mba_data.get("excel_params") or {}
        folder = mba_data.get("folder")
        
        inps_found = False
        if folder:
            from modules.kew.analyse_kew import find_file, parse_inps
            import pandas as pd
            try:
                inps_path = find_file(str(folder), "INPS")
                if inps_path and os.path.isfile(inps_path):
                    # Đọc và ghi dữ liệu INPS bằng bộ parse chuẩn của hệ thống
                    magic, df_raw = parse_inps(inps_path)
                    if df_raw is not None and not df_raw.empty:
                        df_raw.columns = df_raw.columns.str.strip()
                        df, warnings = _mba_extract(df_raw)
                        _mba_write(ws, df)
                        inps_found = True
            except Exception as e:
                print(f"Lỗi đọc file INPS cho {name}: {e}")
                
        if not inps_found:
            # Fallback về logic cũ dùng excel_params nếu không có file INPS
            # --- Đánh giá và ghi vào các ô AE (col 31) ---
            # AE8: Điện áp
            u_max = _parse_float(ep.get("u_max"))
            u_min = _parse_float(ep.get("u_min"))
            u_avg = (u_max + u_min) / 2 if (u_max is not None and u_min is not None) else None
            u_eval, _, _, _ = _eval_voltage(u_max, u_min, u_avg, _MBA_NOMINAL_VOLTAGE_V)
            ws["AE8"] = u_eval
            
            # AE14: Lệch pha áp
            du = _parse_float(ep.get("delta_u"))
            ws["AE14"] = _eval_unbalance(du, du, _V_DEV_LIMIT_PCT)
            
            # AE16: PF
            pf_val = _parse_float(ep.get("cos_phi"))
            if pf_val is not None and abs(pf_val) > 1.0:
                pf_val /= 1000.0
            ws["AE16"] = _eval_pf(pf_val, pf_val, pf_val)
            
            # AE20: THD
            thd = _parse_float(ep.get("thd"))
            ws["AE20"] = _eval_thd([thd], [thd], _THDV_LIMIT_PCT)
            
            # AE23: TDD
            tdd = _parse_float(ep.get("tdd"))
            ws["AE23"] = _eval_thd([tdd], [tdd], _TDD_LIMIT_PCT)

    # 4. Cập nhật bảng tổng hợp tại sheet "Tổn thất MBA"
    if "Tổn thất MBA" in wb.sheetnames:
        ws_ton = wb["Tổn thất MBA"]
        
        # Dòng mẫu A6:J6
        # STT (1) - A6
        # Tên MBA - B6
        # Pdm - D6
        
        for i, mba_data in enumerate(mbas):
            row = 6 + i
            if i > 0:
                # Copy định dạng và công thức từ dòng 6 xuống dòng row
                for col in range(1, 11): # A (1) -> J (10)
                    source_cell = ws_ton.cell(row=6, column=col)
                    new_cell = ws_ton.cell(row=row, column=col)
                    new_cell.value = source_cell.value
                    if source_cell.has_style:
                        from copy import copy
                        new_cell.font = copy(source_cell.font)
                        new_cell.border = copy(source_cell.border)
                        new_cell.fill = copy(source_cell.fill)
                        new_cell.number_format = copy(source_cell.number_format)
                        new_cell.protection = copy(source_cell.protection)
                        new_cell.alignment = copy(source_cell.alignment)

            ep = mba_data.get("excel_params") or {}
            ws_ton.cell(row=row, column=1, value=i + 1) # STT
            name = _nfc(mba_data.get("name") or f"MBA{i+1}")
            ws_ton.cell(row=row, column=2, value=name) # Tên MBA
            pdm = _parse_float(ep.get("pdm"))
            ws_ton.cell(row=row, column=4, value=pdm) # Pdm
            
            # Điền công thức liên kết đến sheet của MBA
            sheet_name = name[:31]
            ws_ton.cell(row=row, column=5, value=f"='{sheet_name}'!AC18")
            ws_ton.cell(row=row, column=6, value=f"='{sheet_name}'!AC16")

    # Lưu kết quả
    wb.save(str(output_path))
    return Path(output_path)
