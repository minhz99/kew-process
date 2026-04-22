import io
import json
import re
from copy import copy

from flask import Blueprint, jsonify, request, send_file

try:
    from openpyxl import load_workbook
except Exception:  # pragma: no cover - dependency guard
    load_workbook = None

excel_bp = Blueprint('excel_bp', __name__)
CELL_ADDR_RE = re.compile(r"^[A-Z]{1,3}[1-9][0-9]*$")


@excel_bp.route('/apply-updates', methods=['POST'])
def apply_updates():
    """Apply value updates to an uploaded .xlsx while preserving existing styles."""
    if load_workbook is None:
        return jsonify({"error": "Thiếu thư viện openpyxl trên server."}), 500

    uploaded_file = request.files.get("file")
    if uploaded_file is None:
        return jsonify({"error": "Cần upload file Excel (.xlsx)."}), 400

    filename = uploaded_file.filename or "KetQua_Excel.xlsx"
    if not filename.lower().endswith(".xlsx"):
        return jsonify({"error": "Chỉ hỗ trợ file định dạng .xlsx."}), 400

    updates_raw = request.form.get("updates", "[]")
    try:
        updates = json.loads(updates_raw)
    except json.JSONDecodeError:
        return jsonify({"error": "Dữ liệu updates không phải JSON hợp lệ."}), 400

    if not isinstance(updates, list):
        return jsonify({"error": "Dữ liệu updates phải là mảng JSON."}), 400

    try:
        workbook = load_workbook(filename=io.BytesIO(uploaded_file.read()))
    except Exception:
        return jsonify({"error": "Không đọc được file Excel đầu vào."}), 400

    for idx, item in enumerate(updates):
        if not isinstance(item, dict):
            return jsonify({"error": f"Update tại vị trí {idx} không hợp lệ."}), 400

        sheet_name = str(item.get("sheet", "")).strip()
        if not sheet_name:
            return jsonify({"error": f"Thiếu tên sheet tại vị trí {idx}."}), 400
        if sheet_name not in workbook.sheetnames:
            return jsonify({"error": f"Không tìm thấy sheet '{sheet_name}' trong file Excel."}), 400

        worksheet = workbook[sheet_name]

        if item.get("type") == "insert_row":
            try:
                row_idx = int(item.get("row"))
            except (ValueError, TypeError):
                return jsonify({"error": f"Dòng chèn tại vị trí {idx} không hợp lệ."}), 400
            
            worksheet.insert_rows(row_idx)
            
            # Copy styles from the row above
            src_row = row_idx - 1
            if src_row > 0:
                for col_idx in range(1, worksheet.max_column + 1):
                    src_cell = worksheet.cell(row=src_row, column=col_idx)
                    tgt_cell = worksheet.cell(row=row_idx, column=col_idx)
                    if src_cell.has_style:
                        tgt_cell.font = copy(src_cell.font)
                        tgt_cell.border = copy(src_cell.border)
                        tgt_cell.fill = copy(src_cell.fill)
                        tgt_cell.number_format = copy(src_cell.number_format)
                        tgt_cell.protection = copy(src_cell.protection)
                        tgt_cell.alignment = copy(src_cell.alignment)
            continue

        cell_address = str(item.get("address", "")).strip().upper()
        if not CELL_ADDR_RE.fullmatch(cell_address):
            return jsonify({"error": f"Địa chỉ ô '{cell_address}' không hợp lệ."}), 400

        worksheet[cell_address].value = item.get("value")

    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)

    output_name = (request.form.get("filename", "") or filename).strip() or "KetQua_Excel.xlsx"
    if not output_name.lower().endswith(".xlsx"):
        output_name = f"{output_name}.xlsx"

    return send_file(
        output,
        as_attachment=True,
        download_name=output_name,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
