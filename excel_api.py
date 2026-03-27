import os
import io
import datetime
import re
import unicodedata
import zipfile
import openpyxl
import copy
from werkzeug.utils import secure_filename
from flask import Blueprint, request, jsonify, send_file

excel_bp = Blueprint('excel_bp', __name__)

BASE_UPLOAD_FOLDER = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'uploads_excel')
os.makedirs(BASE_UPLOAD_FOLDER, exist_ok=True)

def to_number(val):
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return val
    s = str(val).strip()
    if not s:
        return None
    
    if '.' in s and ',' in s:
        if s.rfind('.') > s.rfind(','):
            s = s.replace(',', '')
        else:
            s = s.replace('.', '').replace(',', '.')
    elif '.' in s:
        parts = s.split('.')
        is_thousand_sep = True
        for part in parts[1:]:
            if len(part) != 3:
                is_thousand_sep = False
                break
        if len(parts) > 1 and is_thousand_sep:
            s = s.replace('.', '')
    elif ',' in s:
        parts = s.split(',')
        is_thousand_sep = True
        for part in parts[1:]:
            if len(part) != 3:
                is_thousand_sep = False
                break
        if len(parts) > 1 and is_thousand_sep:
             s = s.replace(',', '')
        else:
             s = s.replace(',', '.')

    try:
        if '.' in s:
            return float(s)
        return int(s)
    except ValueError:
        return val

def fill_excel(ws, pairs, row):
    mapping = [
        ("F", pairs[0][1]), ("G", pairs[0][0]),
        ("I", pairs[1][1]), ("J", pairs[1][0]),
        ("L", pairs[2][1]), ("M", pairs[2][0])
    ]
    for col, value in mapping:
        ws[f"{col}{row}"].value = to_number(value)

def insert_and_setup_row(ws, original_row):
    new_row = original_row + 1
    ws.insert_rows(new_row, 1)
    
    new_merged = []
    for m in ws.merged_cells.ranges:
        min_c, min_r, max_c, max_r = m.min_col, m.min_row, m.max_col, m.max_row
        if min_r < new_row <= max_r:
            max_r += 1
        elif new_row <= min_r:
            min_r += 1
            max_r += 1
        new_merged.append(f"{openpyxl.utils.get_column_letter(min_c)}{min_r}:{openpyxl.utils.get_column_letter(max_c)}{max_r}")
        
    ws.merged_cells.ranges = []
    for m_str in new_merged:
        ws.merge_cells(m_str)

    def shift_complex_match(m):
        prefix, col, mid, row_num = m.group(1), m.group(2), m.group(3), int(m.group(4))
        if row_num >= new_row:
            row_num += 1
        return f"{prefix}{col}{mid}{row_num}"

    for r in range(1, ws.max_row + 1):
        if r == new_row: continue
        for c in range(1, ws.max_column + 1):
            cell = ws.cell(row=r, column=c)
            if cell.data_type == 'f' and cell.value:
                old_val = str(cell.value)
                new_val = re.sub(r'(\$?)([A-Za-z]+)(\$?)(\d+)', shift_complex_match, old_val)
                if new_val != old_val:
                    cell.value = new_val

    def replace_row_in_formula(m):
        prefix, col, mid, row_num = m.group(1), m.group(2), m.group(3), int(m.group(4))
        if row_num == original_row: row_num = new_row
        return f"{prefix}{col}{mid}{row_num}"

    for c in range(1, ws.max_column + 1):
        source_cell = ws.cell(row=original_row, column=c)
        target_cell = ws.cell(row=new_row, column=c)
        
        style_source_cell = source_cell
        if ws.cell(row=original_row, column=4).value is not None:
            style_source_cell = ws.cell(row=new_row + 1, column=c)
            
        if style_source_cell.has_style:
            target_cell._style = copy.copy(style_source_cell._style)
            
        if source_cell.data_type == 'f' and source_cell.value:
            target_cell.value = re.sub(r'(\$?)([A-Za-z]+)(\$?)(\d+)', replace_row_in_formula, str(source_cell.value))
        elif c not in (6, 7, 9, 10, 12, 13):
            if c not in (4, 5) and source_cell.value is not None:
                target_cell.value = source_cell.value

def find_row(ws, target_month, target_period):
    target_month_str = str(target_month).strip()
    target_period_str = str(target_period).strip()
    month_row = None
    for r in range(5, ws.max_row + 1):
        val_d = str(ws.cell(row=r, column=4).value).strip()
        if val_d == target_month_str or val_d == target_month_str + ".0":
            month_row = r
            break
            
    if month_row is None:
        return (target_month - 1) * 4 + 4 + target_period

    for r in range(month_row, ws.max_row + 1):
        val_e = str(ws.cell(row=r, column=5).value).strip()
        if val_e == target_period_str or val_e == target_period_str + ".0":
            return r
        val_d = str(ws.cell(row=r, column=4).value).strip()
        if val_d and val_d != "None" and r != month_row:
            break
        if val_e.lower() == "tổng":
            break
            
    return (target_month - 1) * 4 + 4 + target_period

@excel_bp.route('/upload', methods=['POST'])
def handle_upload():
    session_id = request.form.get('session_id')
    if not session_id:
        return jsonify({'error': 'Thiếu Session ID.'}), 400
        
    if 'file' not in request.files:
        return jsonify({'error': 'Không tìm thấy tệp đính kèm.'}), 400
        
    uploaded_files = request.files.getlist('file')
    if not uploaded_files or uploaded_files[0].filename == '':
        return jsonify({'error': 'Tệp rỗng.'}), 400

    user_folder = os.path.join(BASE_UPLOAD_FOLDER, secure_filename(session_id))
    os.makedirs(user_folder, exist_ok=True)

    saved_files = []
    
    for file in uploaded_files:
        if not file.filename.endswith('.xlsx'):
            continue
            
        filename = secure_filename(file.filename)
        new_path = os.path.join(user_folder, filename)
        
        try:
            file.save(new_path)
            wb = openpyxl.load_workbook(new_path)
            wb.save(new_path)
            saved_files.append(filename)
        except Exception as e:
            return jsonify({'error': f'Lỗi đọc tệp {filename}: {str(e)}'}), 500

    if not saved_files:
        return jsonify({'error': 'Không có file hợp lệ nào được tải lên.'}), 400
        
    return jsonify({'message': f'Tải {len(saved_files)} tệp thành công!', 'filenames': saved_files})

@excel_bp.route('/download', methods=['GET'])
def handle_download():
    session_id = request.args.get('session_id')
    
    if not session_id:
        return "Thiếu thông tin session", 400
        
    user_folder = os.path.join(BASE_UPLOAD_FOLDER, secure_filename(session_id))
    
    if not os.path.exists(user_folder):
        return "Không tìm thấy dữ liệu trên server", 404
        
    files = [f for f in os.listdir(user_folder) if f.endswith('.xlsx')]
    if not files:
        return "Không có file nào để tải về", 404
    
    memory_file = io.BytesIO()
    with zipfile.ZipFile(memory_file, 'w', zipfile.ZIP_DEFLATED) as zf:
        for file in files:
            file_path = os.path.join(user_folder, file)
            zf.write(file_path, arcname=file)
            
    memory_file.seek(0)
    
    timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    dl_name = f"KetQua_Excel_{timestamp}.zip"
    
    return send_file(memory_file, as_attachment=True, download_name=dl_name, mimetype='application/zip')

@excel_bp.route('/submit', methods=['POST'])
def handle_submit():
    data = request.json
    session_id = data.get('session_id')
    filename = data.get('filename')
    
    if not filename or not session_id:
        return jsonify({'error': 'Chưa chọn file Excel hoặc thiếu Session ID.'}), 400

    user_folder = os.path.join(BASE_UPLOAD_FOLDER, secure_filename(session_id))
    file_path = os.path.join(user_folder, secure_filename(filename))
    
    if not os.path.exists(file_path):
        return jsonify({'error': 'File Excel không tồn tại trên server (có thể bạn chưa upload cho phiên làm việc này).'}), 400

    try:
        sheet_name = data.get('sheet_name', '').strip()
        month_str = data.get('month')
        period_str = data.get('period')
        
        if not sheet_name or month_str is None or period_str is None:
            return jsonify({'error': 'Vui lòng nhập đủ tên sheet, tháng và kỳ.'}), 400

        month = int(month_str)
        period = int(period_str)

        if not (1 <= month <= 12): return jsonify({'error': 'Tháng từ 1 đến 12'}), 400
        if period not in (1, 2, 3): return jsonify({'error': 'Kỳ từ 1 đến 3'}), 400

        parsed_groups = []
        mode = data.get('mode')
        
        def format_val(v):
            v = v.strip()
            if v.isdigit() and len(v) >= 4:
                return f"{int(v):,}".replace(",", ".")
            return v

        def strip_accents(s: str) -> str:
            s = unicodedata.normalize("NFD", s)
            return "".join(ch for ch in s if unicodedata.category(ch) != "Mn")

        def extract_pairs_from_text(raw_text: str):
            if not raw_text or not raw_text.strip():
                raise ValueError("Chưa dán dữ liệu vào ô chữ.")

            def classify_line(text: str):
                t = strip_accents(text).lower()
                if "binh thuong" in t: return "bt"
                if "cao diem" in t: return "cd"
                if "thap diem" in t: return "td"
                return None

            num_re = re.compile(r"\d[\d\.]*")
            numeric_items = []
            lines = [ln.strip() for ln in raw_text.replace("\r\n", "\n").replace("\r", "\n").split("\n")]
            for line in lines:
                if not line: continue
                kind = classify_line(line)
                nums = num_re.findall(line)
                if len(nums) >= 2:
                    numeric_items.append((kind, (format_val(nums[0]), format_val(nums[1]))))

            groups = []
            current_group = {"bt": None, "cd": None, "td": None}

            for kind, pair in numeric_items:
                if kind is None:
                    if current_group["bt"] is None: kind = "bt"
                    elif current_group["cd"] is None: kind = "cd"
                    elif current_group["td"] is None: kind = "td"
                    else:
                        groups.append(current_group)
                        current_group = {"bt": None, "cd": None, "td": None}
                        kind = "bt"
                else:
                    if current_group[kind] is not None:
                        groups.append(current_group)
                        current_group = {"bt": None, "cd": None, "td": None}
                current_group[kind] = pair

            if current_group["bt"] or current_group["cd"] or current_group["td"]:
                groups.append(current_group)

            results = []
            for g in groups:
                if any(g[k] is None for k in ["bt", "cd", "td"]):
                    raise ValueError("Không tách được đủ 3 dòng cho 1 kỳ (Bình thường/Cao điểm/Thấp điểm). Bạn hãy dán đầy đủ các dòng.")
                results.append([g["bt"], g["cd"], g["td"]])

            if not results:
                raise ValueError("Không tìm thấy dữ liệu hợp lệ.")

            return results
        
        if mode == 'string':
            raw_data = data.get('raw_data', '').strip()
            try:
                parsed_groups = extract_pairs_from_text(raw_data)
            except ValueError as e:
                return jsonify({'error': str(e)}), 400
                
        elif mode == 'manual':
            b_p, b_u = data.get('bt_price', ''), data.get('bt_usage', '')
            c_p, c_u = data.get('cd_price', ''), data.get('cd_usage', '')
            t_p, t_u = data.get('td_price', ''), data.get('td_usage', '')
            
            if not all([b_p.strip(), b_u.strip(), c_p.strip(), c_u.strip(), t_p.strip(), t_u.strip()]):
                return jsonify({'error': 'Vui lòng điền đủ 6 ô Nhập Từng Ô.'}), 400
            parsed_groups = [[
                (format_val(b_p), format_val(b_u)), 
                (format_val(c_p), format_val(c_u)), 
                (format_val(t_p), format_val(t_u))
            ]]
            
        else:
            return jsonify({'error': 'Chế độ nhập không hợp lệ'}), 400

        wb = openpyxl.load_workbook(file_path)
        if sheet_name not in wb.sheetnames:
            return jsonify({'error': f"Không tìm thấy sheet '{sheet_name}'. Các sheet có sẵn: {', '.join(wb.sheetnames)}"}), 400

        ws = wb[sheet_name]
        
        start_row = find_row(ws, month, period)
        
        inserted_results = []
        for i, group in enumerate(parsed_groups):
            current_write_row = start_row + i
            if i > 0:
                insert_and_setup_row(ws, current_write_row - 1)
            fill_excel(ws, group, current_write_row)
            inserted_results.append({
                'row': current_write_row,
                'parsed_data': group
            })

        wb.save(file_path)

        return jsonify({
            'message': 'Success', 
            'inserted_groups': inserted_results,
            'filename': filename
        })

    except ValueError:
        return jsonify({'error': 'Các giá trị nhập vào không hợp lệ.'}), 400
    except PermissionError:
        return jsonify({'error': 'Không thể ghi file Excel. Hãy chắc chắn bạn không mở file này bằng phần mềm thứ 3.'}), 400
    except Exception as e:
        return jsonify({'error': f'Lỗi hệ thống: {str(e)}'}), 500
